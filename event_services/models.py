from openpyxl import load_workbook
from datetime import datetime, time

def parse_excel_file(file):
    workbook = load_workbook(filename=file)
    sheet = workbook.active
    events = []

    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        date = row_data.get('date')
        title = row_data.get('title')
        time_val = row_data.get('time')

        if isinstance(date, datetime):
            date = date.strftime('%Y-%m-%d')
        if isinstance(time_val, datetime):
            time_val = time_val.strftime('%H:%M')
        elif isinstance(time_val, time):
            time_val = time_val.strftime('%H:%M')

        if date and title and time_val:
            events.append({'date': date, 'title': title, 'time': time_val})
    return events
